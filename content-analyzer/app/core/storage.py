"""
数据存储模块
支持JSON、Excel、CSV格式
"""
import json
import csv
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from app.utils.config import Config
from app.utils.logger import get_logger

logger = get_logger(__name__)

class DataStorage:
    """数据存储类"""
    
    def __init__(self, output_dir: Path = None):
        self.output_dir = output_dir or Config.OUTPUT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def save_json(self, data: List[Dict], filename: str = None) -> Path:
        """
        保存为JSON格式
        
        Args:
            data: 数据列表
            filename: 文件名（不含扩展名）
            
        Returns:
            保存的文件路径
        """
        if filename is None:
            filename = f"crawl_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        filepath = self.output_dir / f"{filename}.json"
        
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            logger.success(f"JSON数据已保存: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"保存JSON失败: {e}")
            raise
    
    def save_excel(self, data: List[Dict], filename: str = None) -> Path:
        """
        保存为Excel格式（带样式）
        
        Args:
            data: 数据列表
            filename: 文件名（不含扩展名）
            
        Returns:
            保存的文件路径
        """
        if filename is None:
            filename = f"crawl_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        filepath = self.output_dir / f"{filename}.xlsx"
        
        try:
            # 创建工作簿
            wb = Workbook()
            
            # 删除默认工作表
            wb.remove(wb.active)
            
            # 创建汇总表
            self._create_summary_sheet(wb, data)
            
            # 创建详细数据表
            self._create_detail_sheet(wb, data)
            
            # 保存
            wb.save(filepath)
            logger.success(f"Excel数据已保存: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"保存Excel失败: {e}")
            raise
    
    def _create_summary_sheet(self, wb: Workbook, data: List[Dict]):
        """创建汇总表"""
        ws = wb.create_sheet("数据汇总", 0)
        
        # 标题样式
        title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='007AFF', end_color='007AFF', fill_type='solid')
        title_align = Alignment(horizontal='center', vertical='center')
        
        # 写入标题
        ws.merge_cells('A1:D1')
        ws['A1'] = '达人账号数据采集报告'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = title_align
        ws.row_dimensions[1].height = 35
        
        # 统计信息
        stats = self._calculate_stats(data)
        
        header_font = Font(name='微软雅黑', size=11, bold=True)
        header_fill = PatternFill(start_color='F5F5F7', end_color='F5F5F7', fill_type='solid')
        
        # 表头
        headers = ['统计项', '数值', '说明', '']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据行
        row = 4
        for key, value in stats.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
    
    def _create_detail_sheet(self, wb: Workbook, data: List[Dict]):
        """创建详细数据表"""
        ws = wb.create_sheet("详细数据")
        
        if not data:
            ws['A1'] = '暂无数据'
            return
        
        # 准备数据
        flat_data = self._flatten_data(data)
        
        if not flat_data:
            ws['A1'] = '数据解析失败'
            return
        
        # 表头样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='007AFF', end_color='007AFF', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        
        # 获取所有字段
        headers = list(flat_data[0].keys())
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # 写入数据
        for row_idx, row_data in enumerate(flat_data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = 20
        
        # 冻结首行
        ws.freeze_panes = 'A2'
    
    def _flatten_data(self, data: List[Dict]) -> List[Dict]:
        """扁平化嵌套数据"""
        flat_list = []
        
        for user in data:
            user_info = {
                '平台': user.get('platform', ''),
                '用户ID': user.get('user_id', ''),
                '昵称': user.get('nickname', ''),
                '简介': user.get('description', ''),
                '粉丝数': user.get('followers', 0),
                '关注数': user.get('following', 0),
                '主页链接': user.get('user_url', ''),
            }
            
            # 处理笔记/视频列表
            contents = user.get('notes', []) or user.get('videos', [])
            
            if contents:
                for content in contents:
                    row = user_info.copy()
                    row.update({
                        '内容标题': content.get('title', ''),
                        '点赞数': content.get('likes', 0),
                        '评论数': content.get('comments', 0),
                        '分享数': content.get('shares', 0),
                        '收藏数': content.get('collects', 0),
                        '播放量': content.get('play_count', 0),
                        '内容链接': content.get('url', ''),
                    })
                    flat_list.append(row)
            else:
                flat_list.append(user_info)
        
        return flat_list
    
    def _calculate_stats(self, data: List[Dict]) -> Dict[str, Any]:
        """计算统计数据"""
        stats = {
            '采集账号总数': len(data),
            '平台分布': '',
            '总内容数': 0,
            '总点赞数': 0,
            '总粉丝数': 0,
            '采集时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
        
        platforms = {}
        total_likes = 0
        total_followers = 0
        total_contents = 0
        
        for user in data:
            # 平台统计
            platform = user.get('platform', 'unknown')
            platforms[platform] = platforms.get(platform, 0) + 1
            
            # 粉丝数
            total_followers += user.get('followers', 0)
            
            # 内容统计
            contents = user.get('notes', []) or user.get('videos', [])
            total_contents += len(contents)
            
            for content in contents:
                total_likes += content.get('likes', 0)
        
        stats['平台分布'] = ', '.join([f"{k}: {v}" for k, v in platforms.items()])
        stats['总内容数'] = total_contents
        stats['总点赞数'] = total_likes
        stats['总粉丝数'] = total_followers
        
        return stats
    
    def save_csv(self, data: List[Dict], filename: str = None) -> Path:
        """
        保存为CSV格式
        
        Args:
            data: 数据列表
            filename: 文件名（不含扩展名）
            
        Returns:
            保存的文件路径
        """
        if filename is None:
            filename = f"crawl_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        filepath = self.output_dir / f"{filename}.csv"
        
        try:
            flat_data = self._flatten_data(data)
            
            if flat_data:
                df = pd.DataFrame(flat_data)
                df.to_csv(filepath, index=False, encoding='utf-8-sig')
                logger.success(f"CSV数据已保存: {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"保存CSV失败: {e}")
            raise
    
    def export(self, data: List[Dict], formats: List[str] = None, filename: str = None) -> Dict[str, Path]:
        """
        导出多种格式
        
        Args:
            data: 数据列表
            formats: 格式列表 ['json', 'excel', 'csv']
            filename: 基础文件名
            
        Returns:
            格式到路径的映射
        """
        if formats is None:
            formats = ['json', 'excel']
        
        results = {}
        
        for fmt in formats:
            try:
                if fmt == 'json':
                    results['json'] = self.save_json(data, filename)
                elif fmt == 'excel':
                    results['excel'] = self.save_excel(data, filename)
                elif fmt == 'csv':
                    results['csv'] = self.save_csv(data, filename)
            except Exception as e:
                logger.error(f"导出 {fmt} 失败: {e}")
        
        return results

class ExcelImporter:
    """Excel导入器"""
    
    def __init__(self):
        self.supported_platforms = {
            'xiaohongshu': ['小红书', 'xiaohongshu', 'xhs'],
            'douyin': ['抖音', 'douyin', 'dy'],
            'weibo': ['微博', 'weibo', 'wb'],
            'bilibili': ['B站', 'bilibili', 'b站', 'blbl']
        }
    
    def import_accounts(self, filepath: Path) -> List[Dict[str, Any]]:
        """
        从Excel导入账号列表
        
        Args:
            filepath: Excel文件路径
            
        Returns:
            账号列表
        """
        logger.info(f"正在导入Excel: {filepath}")
        
        try:
            df = pd.read_excel(filepath)
            accounts = []
            
            for _, row in df.iterrows():
                account = self._parse_row(row)
                if account:
                    accounts.append(account)
            
            logger.success(f"成功导入 {len(accounts)} 个账号")
            return accounts
            
        except Exception as e:
            logger.error(f"导入Excel失败: {e}")
            raise
    
    def _parse_row(self, row: pd.Series) -> Optional[Dict[str, Any]]:
        """解析单行数据"""
        try:
            # 获取平台
            platform = self._detect_platform(row)
            
            # 获取URL
            url = self._extract_url(row)
            
            if not url:
                return None
            
            return {
                'platform': platform,
                'url': url,
                'name': row.get('账号名称', row.get('name', '')),
                'note': row.get('备注', row.get('note', ''))
            }
            
        except Exception as e:
            logger.warning(f"解析行数据失败: {e}")
            return None
    
    def _detect_platform(self, row: pd.Series) -> str:
        """检测平台类型"""
        # 从平台列检测
        platform_col = row.get('平台', row.get('platform', ''))
        if platform_col:
            platform_str = str(platform_col).lower()
            for platform, keywords in self.supported_platforms.items():
                if any(kw in platform_str for kw in keywords):
                    return platform
        
        # 从URL检测
        url = self._extract_url(row)
        if url:
            if 'xiaohongshu' in url or 'xhs' in url:
                return 'xiaohongshu'
            elif 'douyin' in url:
                return 'douyin'
            elif 'weibo' in url:
                return 'weibo'
            elif 'bilibili' in url or 'b23.tv' in url:
                return 'bilibili'
        
        return 'unknown'
    
    def _extract_url(self, row: pd.Series) -> str:
        """提取URL"""
        for col in ['主页链接', '链接', 'url', 'link', '主页']:
            if col in row and pd.notna(row[col]):
                url = str(row[col]).strip()
                if url.startswith('http'):
                    return url
        return ''

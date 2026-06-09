"""
简化版数据存储模块 - 移除pandas依赖
支持JSON、Excel、CSV格式
"""
import json
import csv
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.utils.config import Config
from app.utils.logger import get_logger

logger = get_logger(__name__)


class DataStorage:
    """数据存储类"""
    
    def __init__(self, output_dir: Path = None):
        self.output_dir = output_dir or Config.OUTPUT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def save_json(self, data: List[Dict], filename: str = None) -> Path:
        """保存为JSON格式"""
        if filename is None:
            filename = f"crawl_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        filepath = self.output_dir / f"{filename}.json"
        
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            logger.success(f"JSON数据已保存: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"保存JSON失败: {e}")
            raise
    
    def save_excel(self, data: List[Dict], filename: str = None) -> Path:
        """保存为Excel格式"""
        if filename is None:
            filename = f"crawl_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        filepath = self.output_dir / f"{filename}.xlsx"
        
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            # 创建汇总表
            self._create_summary_sheet(wb, data)
            
            # 创建详细数据表
            self._create_detail_sheet(wb, data)
            
            wb.save(filepath)
            logger.success(f"Excel数据已保存: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"保存Excel失败: {e}")
            raise
    
    def save_csv(self, data: List[Dict], filename: str = None) -> Path:
        """保存为CSV格式"""
        if filename is None:
            filename = f"crawl_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        filepath = self.output_dir / f"{filename}.csv"
        
        try:
            flat_data = self._flatten_data(data)
            if not flat_data:
                logger.warning("无数据可保存")
                return filepath
            
            headers = list(flat_data[0].keys())
            
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(flat_data)
            
            logger.success(f"CSV数据已保存: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"保存CSV失败: {e}")
            raise
    
    def export(self, data: List[Dict], formats: List[str] = None, filename: str = None, suffix: str = None) -> Dict[str, Path]:
        """导出多种格式
        
        Args:
            data: 数据列表
            formats: 导出格式列表
            filename: 基础文件名
            suffix: 文件名后缀（用于中间结果）
        """
        if formats is None:
            formats = ['json', 'excel']
        
        # 如果有后缀，添加到文件名
        if suffix and filename:
            filename = f"{filename}_{suffix}"
        elif suffix and not filename:
            filename = f"crawl_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{suffix}"
        
        results = {}
        
        for fmt in formats:
            try:
                if fmt == 'json':
                    results['json'] = self.save_json(data, filename)
                elif fmt == 'excel':
                    results['excel'] = self.save_excel(data, filename)
                elif fmt == 'csv':
                    results['csv'] = self.save_csv(data, filename)
            except Exception as e:
                logger.error(f"导出 {fmt} 失败: {e}")
        
        return results
    
    def _create_summary_sheet(self, wb: Workbook, data: List[Dict]):
        """创建汇总表"""
        ws = wb.create_sheet("数据汇总", 0)
        
        # 标题样式
        title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='007AFF', end_color='007AFF', fill_type='solid')
        title_align = Alignment(horizontal='center', vertical='center')
        
        # 写入标题
        ws.merge_cells('A1:D1')
        ws['A1'] = '达人账号数据采集报告'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = title_align
        ws.row_dimensions[1].height = 35
        
        # 统计信息
        stats = self._calculate_stats(data)
        
        header_font = Font(name='微软雅黑', size=11, bold=True)
        header_fill = PatternFill(start_color='F5F5F7', end_color='F5F5F7', fill_type='solid')
        
        # 表头
        headers = ['统计项', '数值', '说明', '']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据行
        row = 4
        for key, value in stats.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
    
    def _create_detail_sheet(self, wb: Workbook, data: List[Dict]):
        """创建详细数据表"""
        ws = wb.create_sheet("详细数据")
        
        if not data:
            ws['A1'] = '暂无数据'
            return
        
        flat_data = self._flatten_data(data)
        
        if not flat_data:
            ws['A1'] = '数据解析失败'
            return
        
        # 表头样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='007AFF', end_color='007AFF', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        
        headers = list(flat_data[0].keys())
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # 写入数据
        for row_idx, row_data in enumerate(flat_data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # 调整列宽
        for idx, header in enumerate(headers):
            col_letter = chr(65 + idx) if idx < 26 else 'A' + chr(65 + idx - 26)
            ws.column_dimensions[col_letter].width = 20
        
        # 冻结首行
        ws.freeze_panes = 'A2'
    
    def _flatten_data(self, data: List[Dict]) -> List[Dict]:
        """扁平化嵌套数据"""
        flat_list = []
        
        for user in data:
            user_info = {
                '平台': user.get('platform', ''),
                '用户ID': user.get('user_id', ''),
                '昵称': user.get('nickname', ''),
                '简介': user.get('description', ''),
                '粉丝数': user.get('followers', 0),
                '关注数': user.get('following', 0),
                '主页链接': user.get('user_url', ''),
            }
            
            contents = user.get('notes', []) or user.get('videos', [])
            
            if contents:
                for content in contents:
                    row = user_info.copy()
                    row.update({
                        '内容标题': content.get('title', ''),
                        '发布时间': content.get('publish_date', ''),
                        '点赞数': content.get('interactions', {}).get('likes', 0),
                    })
                    flat_list.append(row)
            else:
                flat_list.append(user_info)
        
        return flat_list
    
    def _calculate_stats(self, data: List[Dict]) -> Dict[str, Any]:
        """计算统计数据"""
        if not data:
            return {"采集账号数": 0, "总内容数": 0}
        
        total_contents = sum(
            len(user.get('notes', []) or user.get('videos', []))
            for user in data
        )
        
        return {
            "采集账号数": len(data),
            "总内容数": total_contents,
            "采集时间": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }


class ExcelImporter:
    """Excel导入器 - 简化版"""
    
    # 支持的列名映射
    COLUMN_MAPPINGS = {
        'platform': ['平台', 'platform', '达人类别', '来源平台'],
        'nickname': ['账号名称', 'nickname', '达人昵称', '账号', '名称', '达人'],
        'url': ['主页链接', 'url', '链接', '主页', '个人主页', '蒲公英链接'],
        'fans': ['粉丝量', 'fans', '粉丝数', 'followers'],
        'note': ['备注', 'note', '说明']
    }
    
    @staticmethod
    def import_accounts(file_path: Path) -> List[Dict[str, Any]]:
        """从Excel导入账号列表"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path)
            ws = wb.active
            
            accounts = []
            headers = [cell.value for cell in ws[1]]
            
            # 检测列名映射
            column_map = ExcelImporter._detect_columns(headers)
            logger.info(f"检测到列映射: {column_map}")
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                
                account = {}
                for idx, header in enumerate(headers):
                    if header and idx < len(row):
                        account[header] = row[idx]
                
                # 标准化字段
                standardized = ExcelImporter._standardize_account(account, column_map)
                
                print(f"[导入调试] 标准化结果: nickname={standardized.get('nickname')}, url={standardized.get('url', '')[:50]}...")
                
                # 只要有昵称和链接就导入
                if standardized.get('nickname') and standardized.get('url'):
                    # 合并原始数据和标准化数据
                    account['platform'] = standardized['platform']
                    account['nickname'] = standardized['nickname']
                    account['url'] = standardized['url']
                    account['fans'] = standardized['fans']
                    account['note'] = standardized['note']
                    account['name'] = standardized['nickname']  # 兼容前端显示
                    accounts.append(account)
                    print(f"[导入调试] ✓ 成功导入账号: {standardized['nickname']}")
                else:
                    print(f"[导入调试] ✗ 跳过账号: 昵称={standardized.get('nickname')}, 有URL={bool(standardized.get('url'))}")
            
            logger.info(f"从Excel导入 {len(accounts)} 个账号")
            return accounts
            
        except Exception as e:
            logger.error(f"导入Excel失败: {e}")
            return []
    
    @staticmethod
    def _detect_columns(headers: List[str]) -> Dict[str, int]:
        """检测列名映射关系"""
        column_map = {}
        
        for idx, header in enumerate(headers):
            if not header:
                continue
            
            header_str = str(header).strip()
            
            for field, possible_names in ExcelImporter.COLUMN_MAPPINGS.items():
                if header_str in possible_names:
                    column_map[field] = idx
                    break
        
        return column_map
    
    @staticmethod
    def _standardize_account(account: Dict, column_map: Dict[str, int]) -> Dict[str, Any]:
        """标准化账号数据"""
        result = {
            'platform': 'xiaohongshu',  # 默认小红书
            'nickname': '',
            'url': '',
            'fans': '',
            'note': ''
        }
        
        headers = list(account.keys())
        
        # 首先优先检测蒲公英链接（不管column_map如何映射）
        pgy_url = None
        for key in account.keys():
            if key and '蒲公英' in str(key):
                pgy_url = str(account[key]).strip()
                print(f"[导入] 找到蒲公英链接: {pgy_url[:60]}...")
                break
        
        # 根据映射提取数据
        for field, idx in column_map.items():
            if idx < len(headers):
                header = headers[idx]
                result[field] = str(account.get(header, '')).strip()
        
        # 如果有蒲公英链接，优先使用蒲公英链接替换掉映射的url
        if pgy_url:
            result['url'] = pgy_url
            print(f"[导入] ✓ 使用蒲公英链接作为采集URL")
        
        # 如果没有检测到映射，尝试直接从原始字段获取
        if not result['nickname']:
            for key in account.keys():
                if key and ('昵称' in str(key) or '名称' in str(key) or '账号' in str(key)):
                    result['nickname'] = str(account[key]).strip()
                    break
        
        # 如果还没有url，再找其他链接
        if not result['url']:
            for key in account.keys():
                if key and ('链接' in str(key) or '主页' in str(key) or 'url' in str(key).lower()):
                    result['url'] = str(account[key]).strip()
                    print(f"[导入] 使用主页链接: {result['url'][:60]}...")
                    break
        
        # 从URL检测平台
        if result['url']:
            url_lower = result['url'].lower()
            if 'xiaohongshu' in url_lower or 'xhs' in url_lower:
                result['platform'] = 'xiaohongshu'
            elif 'douyin' in url_lower:
                result['platform'] = 'douyin'
            elif 'weibo' in url_lower:
                result['platform'] = 'weibo'
            elif 'bilibili' in url_lower:
                result['platform'] = 'bilibili'
        
        return result
